"""
Export Utilities for Threat Reporting System
Provides PDF, CSV, and Excel export functionality for reports and analytics
"""

import io
import csv
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import Report

def export_report_to_pdf(report):
    """
    Export single report to PDF with full details
    
    Args:
        report: Report object from database
    
    Returns:
        BytesIO: PDF file buffer
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom style for title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    title = Paragraph(f"Threat Report #{report.id}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Report Details Table
    report_data = [
        ['Report ID:', f'#{report.id}'],
        ['Category:', report.category.title()],
        ['Problem Type:', report.problem_type],
        ['Status:', report.status.replace('_', ' ').title()],
        ['Submitted:', report.created_at.strftime('%d %B %Y, %I:%M %p')],
        ['Last Updated:', report.updated_at.strftime('%d %B %Y, %I:%M %p')],
    ]
    
    if report.user:
        report_data.append(['Reporter Phone:', report.user.phone])
    
    t = Table(report_data, colWidths=[2*inch, 4*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.3*inch))
    
    # Description Section
    elements.append(Paragraph("Description", heading_style))
    desc_text = report.description.replace('\n', '<br/>')
    elements.append(Paragraph(desc_text, styles['BodyText']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Keywords
    if report.keywords:
        elements.append(Paragraph("Keywords", heading_style))
        keywords_text = ', '.join([kw.keyword for kw in report.keywords])
        elements.append(Paragraph(keywords_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Attachments
    if report.attachments:
        elements.append(Paragraph("Attachments", heading_style))
        for att in report.attachments:
            att_text = f"• {att.filename} ({att.file_type})"
            elements.append(Paragraph(att_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Manager Comments
    if report.manager_comments:
        elements.append(Paragraph("Manager Comments", heading_style))
        comments_text = report.manager_comments.replace('\n', '<br/>')
        elements.append(Paragraph(comments_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Solution
    if report.solutions:
        elements.append(Paragraph("Resolution", heading_style))
        for sol in report.solutions:
            sol_text = sol.solution_text.replace('\n', '<br/>')
            elements.append(Paragraph(sol_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
    
    # Similar Cases
    if report.case_matches:
        elements.append(Paragraph("Similar Police/Government Cases", heading_style))
        cases_data = [['Case Title', 'Similarity', 'Authority']]
        for match in report.case_matches:
            cases_data.append([
                match.past_case.case_title[:30] + '...',
                f"{match.similarity * 100:.0f}%",
                match.past_case.authority
            ])
        cases_table = Table(cases_data, colWidths=[3*inch, 1*inch, 2*inch])
        cases_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ]))
        elements.append(cases_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"<i>Generated on {datetime.now().strftime('%d %B %Y at %I:%M %p')}</i><br/>"
    footer_text += "<i>Threat Reporting System - Confidential Document</i>"
    elements.append(Paragraph(footer_text, styles['Italic']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_reports_to_csv(reports):
    """
    Export multiple reports to CSV format
    
    Args:
        reports: List of Report objects or QuerySet
    
    Returns:
        str: CSV data as string
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    writer.writerow([
        'Report ID',
        'Category',
        'Problem Type',
        'Status',
        'Reporter Phone',
        'Reporter Email',
        'Description',
        'Keywords',
        'Manager Comments',
        'Resolution',
        'Submitted Date',
        'Updated Date',
        'Attachments Count'
    ])
    
    # Data rows
    for report in reports:
        keywords_str = ', '.join([kw.keyword for kw in report.keywords]) if report.keywords else ''
        
        solutions_str = ' | '.join([sol.solution_text for sol in report.solutions]) if report.solutions else ''
        
        attachments_count = len(report.attachments) if report.attachments else 0
        
        writer.writerow([
            report.id,
            report.category,
            report.problem_type,
            report.status,
            report.user.phone if report.user else '',
            report.user.email if report.user and report.user.email else '',
            report.description[:200],  # Truncate for CSV
            keywords_str,
            report.manager_comments or '',
            solutions_str[:500],  # Truncate solutions
            report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '',
            report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if report.updated_at else '',
            attachments_count
        ])
    
    return output.getvalue()


def export_analytics_to_excel(analytics_data):
    """
    Export analytics data to Excel with multiple sheets
    
    Args:
        analytics_data: Dictionary containing analytics data
            {
                'total_reports': int,
                'by_status': {...},
                'by_category': {...},
                'average_resolution_time': float,
                'recent_reports': [...]
            }
    
    Returns:
        BytesIO: Excel file buffer
    """
    output = io.BytesIO()
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # 1. Summary Sheet
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet['A1'] = "Threat Reporting System - Analytics Report"
    summary_sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    summary_sheet['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    summary_sheet.merge_cells('A1:B1')
    
    summary_sheet['A3'] = "Metric"
    summary_sheet['B3'] = "Value"
    for col in ['A3', 'B3']:
        summary_sheet[col].font = Font(bold=True, color="FFFFFF")
        summary_sheet[col].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    
    row = 4
    for key, value in analytics_data.items():
        if not isinstance(value, (list, dict)):
            summary_sheet[f'A{row}'] = key.replace('_', ' ').title()
            summary_sheet[f'B{row}'] = value
            row += 1
    
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 20
    
    # 2. Status Distribution Sheet
    if 'by_status' in analytics_data:
        status_sheet = workbook.create_sheet("By Status")
        status_sheet['A1'] = "Reports by Status"
        status_sheet['A1'].font = Font(size=12, bold=True)
        status_sheet.merge_cells('A1:B1')
        
        status_sheet['A2'] = "Status"
        status_sheet['B2'] = "Count"
        for col in ['A2', 'B2']:
            status_sheet[col].font = Font(bold=True)
            status_sheet[col].fill = PatternFill(start_color="E8EBF5", end_color="E8EBF5", fill_type="solid")
        
        row = 3
        for status, count in analytics_data['by_status'].items():
            status_sheet[f'A{row}'] = status.replace('_', ' ').title()
            status_sheet[f'B{row}'] = count
            row += 1
        
        status_sheet.column_dimensions['A'].width = 25
        status_sheet.column_dimensions['B'].width = 15
    
    # Save workbook
    workbook.save(output)
    output.seek(0)
    return output


def generate_csv_headers():
    """Get CSV export headers"""
    return [
        'Report ID',
        'Category',
        'Problem Type',
        'Status',
        'Reporter Phone',
        'Reporter Email',
        'Description',
        'Keywords',
        'Manager Comments',
        'Resolution',
        'Submitted Date',
        'Updated Date',
        'Attachments Count'
    ]
